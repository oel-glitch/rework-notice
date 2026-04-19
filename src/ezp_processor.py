"""
Модуль обработки ЭЗП (Электронных Заказных Писем)
Парсинг адресов из PDF, нормализация по ГОСТ, заполнение Excel
"""

import re
import os
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import logging
import fitz  # PyMuPDF


@dataclass
class Recipient:
    """Одна запись адресата, извлечённая из ЭЗП-PDF."""
    recipient_name: str
    address_raw: str
    entity_type: int  # 1 — юрлицо, 0 — физлицо
    emails: List[str] = field(default_factory=list)
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class AddressParser:
    """Парсер адресов из PDF документов (координатный разбор блока адресата)."""

    # Минимальный x0 для столбца адресата. Шапка ОАТИ идёт от x≈260,
    # блок адресата — от x≈305+. Порог 290 даёт чистое разделение.
    ADDRESSEE_MIN_X = 290.0

    # Горизонтальная полоса поиска: ниже шапки ОАТИ (≈60pt сверху),
    # но выше тела письма (до 55% высоты страницы).
    HEADER_SKIP_TOP = 60.0
    CLIP_BOTTOM_RATIO = 0.55

    # Допуск на группировку спанов в одну строку по y0 (пункты).
    LINE_Y_TOLERANCE = 3.0

    # Максимальный зазор между строками внутри блока адресата, пункты.
    # Наблюдения: внутри блока бывают зазоры до 30pt (между преамбулой/ФИО/адресом
    # в примере с депутатом), между блоком и телом письма — от 37pt. Порог 33
    # даёт чистую отсечку.
    MAX_BLOCK_GAP = 33.0

    # Маркеры, однозначно указывающие на строку адреса.
    # Используем \b для точных слов: "города Москвы" не должно ловиться как "г.".
    ADDRESS_MARKERS = re.compile(
        r"(?:\b(?:ул|просп|пер|пр-т|пр-д|ш|наб|пл|бул|б-р|д|корп|стр|кв|оф|пом|г|обл|р-н|пос|дер|с)\.|"
        r"\bиндекс\b|"
        r"\b\d{6}\b)",
        re.IGNORECASE,
    )

    # ФИО с инициалами: "Прописновой Е.О.", "Иванов И. И." и т.п.
    NAME_PATTERN = re.compile(
        r"^\s*[А-ЯЁ][а-яё]+(?:-[А-ЯЁ][а-яё]+)?\s+[А-ЯЁ]\.\s*[А-ЯЁ]?\.?\s*$"
    )

    EMAIL_PATTERN = re.compile(r"\S+@\S+\.\S+")

    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.AddressParser")
        self._entity_detector = EntityTypeDetector()

    # ---- Публичный API ---------------------------------------------------

    def extract_recipients(self, pdf_path: str) -> List[Recipient]:
        """
        Извлекает список адресатов из PDF (первая страница).

        Логика:
        1. Берём все span'ы в правом столбце (x0 > ADDRESSEE_MIN_X), ниже шапки.
        2. Группируем в строки по y0.
        3. Сегментируем на группы по вертикальным зазорам > RECIPIENT_GAP_THRESHOLD.
        4. В каждой группе классифицируем строки: email / address / name / other.
        5. Собираем Recipient.
        """
        try:
            doc = fitz.open(pdf_path)
        except Exception as e:
            self.logger.error(f"Не удалось открыть PDF {pdf_path}: {e}")
            return []

        try:
            if doc.page_count == 0:
                return []
            page = doc[0]
            page_w, page_h = page.rect.width, page.rect.height
            clip = fitz.Rect(
                self.ADDRESSEE_MIN_X - 5,  # небольшой запас слева
                self.HEADER_SKIP_TOP,
                page_w,
                page_h * self.CLIP_BOTTOM_RATIO,
            )
            raw = page.get_text("dict", clip=clip)
        finally:
            doc.close()

        spans = self._collect_spans(raw)
        if not spans:
            self.logger.warning(f"В блоке адресата нет текста: {pdf_path}")
            return []

        lines = self._group_spans_into_lines(spans)
        groups = self._segment_into_recipients(lines)

        recipients: List[Recipient] = []
        for group in groups:
            rec = self._build_recipient(group)
            if rec is not None:
                recipients.append(rec)

        if not recipients:
            self.logger.warning(f"Адресаты не распознаны в {pdf_path}")
        else:
            self.logger.info(
                f"Извлечено адресатов из {pdf_path}: {len(recipients)}"
            )
        return recipients

    def extract_address_from_pdf(self, pdf_path: str) -> Optional[str]:
        """
        Обратная совместимость: возвращает "сырой" адрес первого адресата
        (recipient_name + address_raw), как раньше ожидал старый процессор.
        """
        recipients = self.extract_recipients(pdf_path)
        if not recipients:
            return None
        r = recipients[0]
        parts = [p for p in (r.recipient_name, r.address_raw) if p]
        return ", ".join(parts) if parts else None

    # ---- Внутреннее -----------------------------------------------------

    def _collect_spans(self, raw: dict) -> List[dict]:
        """Плоский список span'ов только из правого столбца."""
        out: List[dict] = []
        for block in raw.get("blocks", []):
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    text = (span.get("text") or "").strip()
                    if not text:
                        continue
                    bbox = span.get("bbox") or (0, 0, 0, 0)
                    x0, y0, x1, y1 = bbox
                    if x0 < self.ADDRESSEE_MIN_X:
                        continue
                    out.append({
                        "text": text,
                        "x0": x0, "y0": y0, "x1": x1, "y1": y1,
                        "bold": bool(span.get("flags", 0) & 16),
                        "size": span.get("size", 0),
                    })
        return out

    def _group_spans_into_lines(self, spans: List[dict]) -> List[dict]:
        """Группирует spans в строки по y0 (±LINE_Y_TOLERANCE)."""
        spans_sorted = sorted(spans, key=lambda s: (s["y0"], s["x0"]))
        lines: List[dict] = []
        for s in spans_sorted:
            if lines and abs(s["y0"] - lines[-1]["y0"]) <= self.LINE_Y_TOLERANCE:
                line = lines[-1]
                line["parts"].append(s)
                line["x1"] = max(line["x1"], s["x1"])
                line["y1"] = max(line["y1"], s["y1"])
                line["bold"] = line["bold"] and s["bold"]
            else:
                lines.append({
                    "y0": s["y0"], "y1": s["y1"],
                    "x0": s["x0"], "x1": s["x1"],
                    "bold": s["bold"],
                    "parts": [s],
                })
        # Собираем текст строки
        for line in lines:
            parts = sorted(line["parts"], key=lambda p: p["x0"])
            line["text"] = " ".join(p["text"] for p in parts).strip()
        return lines

    def _segment_into_recipients(self, lines: List[dict]) -> List[List[dict]]:
        """
        Возвращает один блок адресата: все строки до первого большого вертикального
        зазора (> MAX_BLOCK_GAP). Всё, что ниже — это уже тело письма.

        Поддержка нескольких адресатов в одном PDF отложена — плагина пока нет
        реальных образцов, и текущая эвристика "первый непрерывный блок" надёжнее
        в production, чем попытка угадать multi-recipient по зазорам.
        """
        if not lines:
            return []
        block: List[dict] = [lines[0]]
        for prev, curr in zip(lines, lines[1:]):
            gap = curr["y0"] - prev["y1"]
            if gap > self.MAX_BLOCK_GAP:
                break
            block.append(curr)
        return [block]

    def _classify_line(self, text: str) -> str:
        """Возвращает тип строки: 'email' | 'address' | 'name' | 'other'."""
        if self.EMAIL_PATTERN.search(text):
            return "email"
        if self.ADDRESS_MARKERS.search(text):
            return "address"
        if self.NAME_PATTERN.match(text):
            return "name"
        return "other"

    def _build_recipient(self, group: List[dict]) -> Optional[Recipient]:
        """Собирает Recipient из группы строк."""
        name_lines: List[str] = []
        org_lines: List[str] = []
        address_lines: List[str] = []
        emails: List[str] = []

        for line in group:
            text = line["text"]
            kind = self._classify_line(text)
            if kind == "email":
                for m in self.EMAIL_PATTERN.findall(text):
                    emails.append(m)
            elif kind == "address":
                address_lines.append(text)
            elif kind == "name":
                name_lines.append(text)
            else:
                org_lines.append(text)

        if not (name_lines or org_lines or address_lines):
            return None
        # Реальный адресат обязан иметь строку адреса. Группа без адреса —
        # почти всегда обрывок тела письма, попавший в правую колонку.
        if not address_lines:
            self.logger.debug("Группа без адресных маркеров — пропускаем")
            return None

        # Формируем имя получателя
        if name_lines:
            name = " ".join(name_lines).strip()
            preamble = " ".join(org_lines).strip()
            if preamble:
                recipient_name = f"{preamble} {name}".strip()
                # Длинная преамбула ("Депутату ...", "Директору ...") → юрлицо
                entity_type = 1
            else:
                recipient_name = name
                entity_type = 0
        elif org_lines:
            recipient_name = " ".join(org_lines).strip()
            entity_type = self._entity_detector.detect_entity_type(recipient_name)
            # Для чистой организации без маркеров (редко) — всё равно юрлицо,
            # т.к. в ЭЗП адресатом-физлицом без имени письмо не шлют.
            if entity_type == 0:
                entity_type = 1
        else:
            # Только адрес без имени — пропускаем
            self.logger.warning("Группа содержит только адрес без получателя")
            return None

        address_raw = ", ".join(address_lines).strip()

        return Recipient(
            recipient_name=recipient_name,
            address_raw=address_raw,
            entity_type=entity_type,
            emails=emails,
        )
    
    # Регистрационный штамп ОАТИ: "Документ зарегистрирован № 01-13-1492/26 от 08.03.2026 ..."
    # Штамп штампуется на PDF при регистрации исходящего, находится обычно внизу страницы.
    # Поддерживаемые форматы номера: 01-13-1492/26, 01-21-П-7779/25-1, 01-16.5-584/26.
    _DOC_CORE = r"\d{2}-\d{1,2}(?:\.\d+)?-[\wА-ЯЁа-яё-]+/\d{2}(?:-\d+)?"
    REG_STAMP_PATTERN = re.compile(
        r"Документ\s+зарегистрирован\s*№\s*(" + _DOC_CORE + r")"
    )
    # Запасной паттерн: любой текст вида "01-XX-..." формата ОАТИ
    DOC_NUMBER_FALLBACK = re.compile(r"\b" + _DOC_CORE + r"\b")

    def extract_document_number_from_pdf(self, pdf_path: str) -> Optional[str]:
        """
        Извлекает регистрационный номер документа из PDF.

        Сначала ищет регистрационный штамп ОАТИ "Документ зарегистрирован № ..."
        (появляется при регистрации исходящего письма, штампуется обычно внизу
        первой страницы). Как запасной вариант — ищет шаблон "XX-XX-.../YY"
        в верхней части страницы (на случай если письмо ещё не зарегистрировано).
        """
        try:
            doc = fitz.open(pdf_path)
        except Exception as e:
            self.logger.error(f"Не удалось открыть PDF {pdf_path}: {e}")
            return None

        try:
            if doc.page_count == 0:
                return None
            page = doc[0]
            full_text = page.get_text("text")
        finally:
            doc.close()

        # 1. Регистрационный штамп — самый надёжный источник
        m = self.REG_STAMP_PATTERN.search(full_text)
        if m:
            doc_number = m.group(1).strip()
            self.logger.info(f"Номер документа (штамп): {pdf_path} → {doc_number}")
            return doc_number

        # 2. Fallback: любое совпадение по шаблону в тексте страницы
        m = self.DOC_NUMBER_FALLBACK.search(full_text)
        if m:
            doc_number = m.group(0).strip()
            self.logger.info(f"Номер документа (fallback): {pdf_path} → {doc_number}")
            return doc_number

        self.logger.warning(f"Номер документа не найден в {pdf_path}")
        return None


class AddressNormalizer:
    """
    Нормализует адрес в формат "ИНДЕКС, улица/дом/корпус/..., г. Город".

    Вход — сырой адрес в одну строку (из AddressParser.address_raw), уже
    содержащий стандартные сокращения (ул., д., корп., стр., г.) — они
    берутся из шапки адресата PDF. Задача нормализатора — вытащить индекс
    и город, почистить мусор (лишние пробелы и запятые) и собрать в эталонном
    порядке: сначала индекс, затем street/house-блок, затем "г. Город".
    """

    # 6-значный индекс как отдельный токен
    _INDEX_RE = re.compile(r"\b(\d{6})\b")

    # "г. Название" — город (допускаем "Новый Арбат"-подобное — одно-два слова)
    _CITY_RE = re.compile(r"г\.\s*[А-ЯЁ][а-яё]+(?:[-\s][А-ЯЁ][а-яё]+)?")

    # Полные слова → стандартные сокращения. Паттерны специально с границей
    # справа (пробел/запятая/конец строки), чтобы не трогать уже сокращённое.
    _FULLFORM_REPLACEMENTS = [
        (re.compile(r"\bулица\b", re.IGNORECASE), "ул."),
        (re.compile(r"\bпроспект\b", re.IGNORECASE), "просп."),
        (re.compile(r"\bпереулок\b", re.IGNORECASE), "пер."),
        (re.compile(r"\bбульвар\b", re.IGNORECASE), "б-р"),
        (re.compile(r"\bпроезд\b", re.IGNORECASE), "пр-д"),
        (re.compile(r"\bшоссе\b", re.IGNORECASE), "ш."),
        (re.compile(r"\bнабережная\b", re.IGNORECASE), "наб."),
        (re.compile(r"\bплощадь\b", re.IGNORECASE), "пл."),
        (re.compile(r"\bдом\b", re.IGNORECASE), "д."),
        (re.compile(r"\bкорпус\b", re.IGNORECASE), "корп."),
        (re.compile(r"\bстроение\b", re.IGNORECASE), "стр."),
        (re.compile(r"\bквартира\b", re.IGNORECASE), "кв."),
        (re.compile(r"\bофис\b", re.IGNORECASE), "оф."),
        (re.compile(r"\bпомещение\b", re.IGNORECASE), "пом."),
        (re.compile(r"\bгород\b(?!а)", re.IGNORECASE), "г."),  # "город" но не "города"
        (re.compile(r"\bобласть\b", re.IGNORECASE), "обл."),
        (re.compile(r"\bрайон\b", re.IGNORECASE), "р-н"),
    ]

    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.AddressNormalizer")

    def normalize_address(self, address: str) -> str:
        """Приводит адрес к виду "123007, 1-я Магистральная ул., д. 17, стр. 4, г. Москва"."""
        if not address:
            return ""

        text = address

        # 1. Заменяем полные слова на сокращения (только там, где их нет).
        for pattern, repl in self._FULLFORM_REPLACEMENTS:
            text = pattern.sub(repl, text)

        # 2. Извлекаем индекс и убираем его из тела.
        index = ""
        m = self._INDEX_RE.search(text)
        if m:
            index = m.group(1)
            text = text[: m.start()] + text[m.end():]

        # 3. Извлекаем "г. Город" и убираем из тела.
        city = ""
        m = self._CITY_RE.search(text)
        if m:
            city = m.group(0).strip()
            text = text[: m.start()] + text[m.end():]

        # 4. Чистим остаток: нормализуем пробелы, схлопываем повторные запятые,
        #    убираем висячие разделители по краям.
        remainder = self._tidy(text)

        # 5. Собираем в эталонном порядке.
        parts = [p for p in (index, remainder, city) if p]
        result = ", ".join(parts)
        self.logger.debug(f"Нормализация: {address!r} → {result!r}")
        return result

    @staticmethod
    def _tidy(text: str) -> str:
        """Схлопывает повторные запятые/пробелы и срезает их по краям."""
        text = re.sub(r"\s+", " ", text)
        text = re.sub(r"(\s*,\s*){2,}", ", ", text)  # ", ,," → ", "
        text = text.strip(" ,")
        return text


class EntityTypeDetector:
    """Определение типа получателя (юридическое/физическое лицо)"""
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.EntityTypeDetector")
        
        # Признаки юридического лица
        self.legal_entity_prefixes = [
            'ООО', 'ОАО', 'ЗАО', 'ПАО', 'АО',
            'ИП', 'ИП ', 'Индивидуальный предприниматель',
            'ГУП', 'МУП', 'ФГУП',
            'НКО', 'АНО', 'Фонд',
            'Товарищество', 'Кооператив',
            'Учреждение', 'Предприятие', 'Организация',
            'Компания', 'Корпорация', 'Холдинг',
            'Банк', 'Страховая'
        ]
        
        # Признаки государственных органов
        self.government_keywords = [
            'Правительство', 'Министерство', 'Департамент',
            'Управление', 'Префектура', 'Администрация',
            'Инспекция', 'Служба', 'Агентство',
            'Комитет', 'Комиссия', 'Совет'
        ]
    
    def detect_entity_type(self, text: str) -> int:
        """
        Определяет тип получателя по тексту
        
        Args:
            text: Текст для анализа (адрес или название)
            
        Returns:
            1 - юридическое лицо, 0 - физическое лицо
        """
        if not text:
            return 0
        
        text_upper = text.upper()
        
        # Проверяем признаки юридического лица
        for prefix in self.legal_entity_prefixes:
            if prefix.upper() in text_upper:
                self.logger.debug(f"Найден признак юр. лица: {prefix} в '{text}'")
                return 1
        
        # Проверяем признаки государственного органа
        for keyword in self.government_keywords:
            if keyword.upper() in text_upper:
                self.logger.debug(f"Найден признак гос. органа: {keyword} в '{text}'")
                return 1
        
        # Проверяем наличие кавычек (часто указывает на название организации)
        if '"' in text or '«' in text or '»' in text:
            self.logger.debug(f"Найдены кавычки (возможно организация) в '{text}'")
            return 1
        
        # По умолчанию считаем физическим лицом
        self.logger.debug(f"Определено как физ. лицо: '{text}'")
        return 0


class EZPProcessor:
    """Основной класс обработки ЭЗП"""
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.EZPProcessor")
        self.address_parser = AddressParser()
        self.address_normalizer = AddressNormalizer()
        self.entity_detector = EntityTypeDetector()
        
    def create_template_backup(self, excel_path: str, log_callback=None) -> Optional[str]:
        """
        Создает резервную копию шаблона при первой обработке
        Проверяет наличие существующих резервных копий на диске
        
        Args:
            excel_path: Путь к оригинальному Excel файлу
            log_callback: Функция для вывода логов
            
        Returns:
            Путь к резервной копии или None если уже была создана ранее
        """
        try:
            import glob
            from datetime import datetime
            import shutil
            
            # Проверяем, существует ли исходный файл
            if not os.path.exists(excel_path):
                error_msg = f"Шаблон не найден: {excel_path}"
                self.logger.error(error_msg)
                if log_callback:
                    log_callback(f"✗ ОШИБКА: {error_msg}")
                return None
            
            # Формируем путь и паттерн для поиска существующих копий
            base_path = Path(excel_path)
            backup_dir = base_path.parent
            backup_pattern = str(backup_dir / f"{base_path.stem}_backup_*{base_path.suffix}")
            
            # Проверяем, есть ли уже резервные копии на диске
            existing_backups = glob.glob(backup_pattern)
            if existing_backups:
                # Сортируем по времени модификации, берем самую свежую
                existing_backups.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                latest_backup = existing_backups[0]
                backup_name = os.path.basename(latest_backup)
                msg = f"✓ Резервная копия уже существует: {backup_name}"
                self.logger.info(msg)
                if log_callback:
                    log_callback(msg)
                    if len(existing_backups) > 1:
                        log_callback(f"  (Всего найдено резервных копий: {len(existing_backups)})")
                return None
            
            # Создаем новую резервную копию
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_name = f"{base_path.stem}_backup_{timestamp}{base_path.suffix}"
            backup_path = backup_dir / backup_name
            
            # Копируем файл
            shutil.copy2(excel_path, backup_path)
            
            msg = f"✓ Создана резервная копия шаблона: {backup_name}"
            self.logger.info(msg)
            if log_callback:
                log_callback(msg)
                log_callback("Рекомендация: Сохраните эту копию для будущего использования")
                
            return str(backup_path)
            
        except Exception as e:
            error_msg = f"Не удалось создать резервную копию: {e}"
            self.logger.error(error_msg)
            if log_callback:
                log_callback(f"⚠ {error_msg}")
            return None
        
    def process_ezp(self, pdf_files: List[str], output_filename: Optional[str] = None, progress_callback=None, log_callback=None) -> Tuple[bool, str, Dict]:
        """
        Обрабатывает ЭЗП: парсит PDF, нормализует адреса, заполняет Excel
        
        НОВАЯ ЛОГИКА v4.3:
        - Принимает список PDF файлов напрямую (не папку!)
        - Автоматически использует шаблон templates/ezp_template.xls
        - Создает новый файл в output/ с уникальным именем
        - Заполняет 13 столбцов согласно структуре Почты России
        - Работает OFFLINE без AI (fallback классификация)
        
        Args:
            pdf_files: Список путей к PDF файлам для обработки
            output_filename: Опциональное имя выходного файла (если None, генерируется автоматически)
            progress_callback: Функция для обновления прогресса (0-100)
            log_callback: Функция для вывода логов
            
        Returns:
            (успех, сообщение, статистика)
        """
        stats = {
            'total_pdfs': 0,
            'processed': 0,
            'failed': 0,
            'addresses_found': 0,
            'document_numbers_found': 0,
            'legal_entities': 0,
            'physical_persons': 0
        }
        
        # КОНСТАНТЫ для заполнения таблицы
        SENDER_ADDRESS = "119019, г Москва, ул Новый Арбат, д 21"
        MAIL_CATEGORY = 1
        MAIL_RANK = 1
        
        try:
            # Проверяем доступность библиотек
            if not PANDAS_AVAILABLE:
                return False, "Библиотека pandas не установлена", stats
            if not OPENPYXL_AVAILABLE:
                return False, "Библиотека openpyxl не установлена", stats
            
            if log_callback:
                log_callback("="*60)
                log_callback("НАЧАЛО ОБРАБОТКИ ЭЗП")
                log_callback("="*60)
            
            # Определяем пути к шаблону и выходному файлу.
            # Предпочитаем .xlsx (нужен openpyxl для форматирования результата);
            # если есть только старый .xls — используем его (pd.read_excel сам подхватит).
            template_xlsx = Path("templates/ezp_template.xlsx")
            template_xls = Path("templates/ezp_template.xls")
            if template_xlsx.exists():
                template_path = template_xlsx
            elif template_xls.exists():
                template_path = template_xls
            else:
                error_msg = (
                    "Шаблон не найден: ожидался templates/ezp_template.xlsx "
                    "или templates/ezp_template.xls"
                )
                if log_callback:
                    log_callback(f"✗ ОШИБКА: {error_msg}")
                return False, error_msg, stats
            
            # Создаем output директорию если не существует
            output_dir = Path("output")
            output_dir.mkdir(exist_ok=True)
            
            # Генерируем имя выходного файла
            if output_filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_filename = f"Отправка_{timestamp}.xlsx"
            
            excel_path = str(output_dir / output_filename)
            
            # Копируем шаблон в output
            if log_callback:
                log_callback(f"📋 Использую шаблон: {template_path}")
                log_callback(f"📁 Создаю файл: {excel_path}")
            
            shutil.copy2(template_path, excel_path)
            
            # Загружаем скопированный файл
            if log_callback:
                log_callback(f"Загрузка шаблона...")
            
            df = pd.read_excel(excel_path, engine='openpyxl')
            
            if log_callback:
                log_callback(f"Шаблон загружен: {len(df)} строк, {len(df.columns)} столбцов")
                log_callback(f"Столбцы: {list(df.columns)}")
            
            # Очищаем старые данные (оставляем только заголовки)
            df = df.iloc[:0].copy()  # Удаляем все строки, оставляем структуру
            
            if log_callback:
                log_callback("✓ Старые данные очищены, заголовки сохранены")
            
            # Используем переданный список PDF файлов
            stats['total_pdfs'] = len(pdf_files)
            
            if not pdf_files:
                return False, "Не переданы PDF файлы для обработки", stats
            
            if log_callback:
                log_callback(f"Файлов к обработке: {len(pdf_files)}")
                log_callback("="*60)
            
            # Обрабатываем каждый PDF
            total = len(pdf_files)
            row_counter = 0  # Сквозная нумерация FILE_NAME (письма, не PDF)
            for idx, pdf_path in enumerate(pdf_files):
                try:
                    filename = os.path.basename(pdf_path)

                    if log_callback:
                        log_callback(f"[{idx + 1}/{total}] Обработка: {filename}")

                    # Новый API: список адресатов из одного PDF
                    recipients = self.address_parser.extract_recipients(pdf_path)

                    if not recipients:
                        if log_callback:
                            log_callback(f"  ✗ Адресаты не найдены")
                        stats['failed'] += 1
                        continue

                    # Извлекаем номер документа (левый верхний угол) — общий для всех адресатов PDF
                    doc_number = self.address_parser.extract_document_number_from_pdf(pdf_path)
                    if doc_number:
                        stats['document_numbers_found'] += 1
                    else:
                        doc_number = ""

                    for rec in recipients:
                        row_counter += 1
                        stats['addresses_found'] += 1

                        # Нормализуем адрес по ГОСТ через существующий нормализатор
                        address_normalized = self.address_normalizer.normalize_address(rec.address_raw)

                        entity_type = rec.entity_type
                        if entity_type == 1:
                            stats['legal_entities'] += 1
                        else:
                            stats['physical_persons'] += 1

                        # Формируем строку данных для таблицы (13 столбцов)
                        row_data = {
                            'FILE_NAME': row_counter,
                            'ADDRESSLINE_TO': address_normalized,
                            'RECIPIENT_TYPE': entity_type,
                            'RECIPIENT': rec.recipient_name,
                            'INN': None,
                            'KPP': None,
                            'LETTER_REG_NUMBER': doc_number,
                            'LETTER_TITLE': doc_number,
                            'MAILCATEGORY': MAIL_CATEGORY,
                            'ADDRESSLINE_RETURN': SENDER_ADDRESS,
                            'WOMAILRANK': MAIL_RANK,
                            'ADDITIONAL_INFO': None,
                            'LETTER_COMMENT': doc_number,
                        }

                        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

                        if log_callback:
                            log_callback(f"  ✓ Строка {row_counter}: {rec.recipient_name[:50]}")
                            log_callback(f"    Адрес: {address_normalized[:60]}")
                            log_callback(f"    Тип: {'Юрлицо' if entity_type == 1 else 'Физлицо'}")

                    stats['processed'] += 1

                    # Обновляем прогресс
                    if progress_callback:
                        progress = int((idx + 1) / total * 100)
                        progress_callback(progress)
                        
                except Exception as e:
                    if log_callback:
                        log_callback(f"  ✗ Ошибка обработки {filename}: {e}")
                    stats['failed'] += 1
            
            # Сохраняем в ТОТ ЖЕ файл (перезаписываем шаблон)
            if log_callback:
                log_callback("="*60)
                log_callback("Сохранение результата...")
            
            # Проверяем доступность openpyxl
            if not OPENPYXL_AVAILABLE:
                error_msg = "Библиотека openpyxl не установлена. Установите: pip install openpyxl"
                if log_callback:
                    log_callback(f"✗ ОШИБКА: {error_msg}")
                return False, error_msg, stats
            
            # Загружаем существующий .xlsx workbook для сохранения форматирования
            from copy import copy as style_copy
            
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Определяем количество строк в оригинальном файле
            original_rows = ws.max_row
            
            # Константы
            NUM_COLS = 13
            col_names = ['FILE_NAME', 'ADDRESSLINE_TO', 'RECIPIENT_TYPE', 'RECIPIENT', 
                        'INN', 'KPP', 'LETTER_REG_NUMBER', 'LETTER_TITLE', 
                        'MAILCATEGORY', 'ADDRESSLINE_RETURN', 'WOMAILRANK', 
                        'ADDITIONAL_INFO', 'LETTER_COMMENT']
            
            # Кэшируем стили из template row (строка 2 - первая строка данных)
            template_row = 2 if ws.max_row >= 2 else 1
            style_cache = []
            for c in range(1, NUM_COLS + 1):
                tpl = ws.cell(row=template_row, column=c)
                style_cache.append({
                    'font': style_copy(tpl.font),
                    'fill': style_copy(tpl.fill),
                    'border': style_copy(tpl.border),
                    'alignment': style_copy(tpl.alignment),
                    'numfmt': tpl.number_format,
                    'protection': style_copy(tpl.protection)
                })
            
            # Сохраняем высоту template row для копирования в новые строки
            template_row_height = ws.row_dimensions[template_row].height
            
            # Очищаем старые данные (строки 2+), сохраняя форматирование И формулы
            for row in ws.iter_rows(min_row=2, max_row=original_rows, max_col=NUM_COLS):
                for cell in row:
                    # КРИТИЧНО: НЕ трогаем формулы! Проверяем все возможные представления формул
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                        continue  # Пропускаем ячейки с формулами
                    cell.value = None  # Очищаем только обычные ячейки
            
            # Записываем новые данные с применением стилей из template
            for offset, row_data in enumerate(df.to_dict('records')):
                r = 2 + offset  # Строка 2 = первая строка данных (после заголовка)
                
                # Для новых строк (за пределами original template) копируем высоту
                if r > original_rows and template_row_height:
                    ws.row_dimensions[r].height = template_row_height
                
                for c, col_name in enumerate(col_names, start=1):
                    cell = ws.cell(row=r, column=c)
                    
                    # Проверяем наличие формулы - НЕ перезаписываем формулы (все представления)
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                        continue  # Пропускаем ячейки с формулами
                    
                    cache = style_cache[c - 1]
                    
                    # Применяем кэшированные стили из template
                    cell.font = cache['font']
                    cell.fill = cache['fill']
                    cell.border = cache['border']
                    cell.alignment = cache['alignment']
                    cell.number_format = cache['numfmt']
                    cell.protection = cache['protection']
                    
                    # Устанавливаем значение (None для пустых полей, не пустую строку!)
                    value = row_data.get(col_name)
                    # КРИТИЧНО: pandas NaN и пустые строки → None для truly blank cells
                    if pd.isna(value) or value == '':
                        value = None
                    cell.value = value  # None остаётся None, пустые ячейки остаются пустыми
            
            # Расширяем границы печати до последней заполненной строки
            if len(df) > 0:
                last_row = len(df) + 1  # +1 для заголовка
                print_area = f"A1:M{last_row}"
                ws.print_area = print_area
                
                if log_callback:
                    log_callback(f"✓ Границы печати установлены: {print_area}")
            
            # Сохраняем
            wb.save(excel_path)
            
            if log_callback:
                log_callback("✓ Файл сохранён в формате .xlsx с сохранением форматирования")
            
            if log_callback:
                log_callback(f"✓ Результат сохранён в исходный файл: {os.path.basename(excel_path)}")
                log_callback("="*60)
                log_callback("СТАТИСТИКА:")
                log_callback(f"  Всего PDF файлов: {stats['total_pdfs']}")
                log_callback(f"  Успешно обработано: {stats['processed']}")
                log_callback(f"  Ошибок: {stats['failed']}")
                log_callback(f"  Адресов найдено: {stats['addresses_found']}")
                log_callback(f"  Номеров документов найдено: {stats['document_numbers_found']}")
                log_callback(f"  Юрлиц: {stats['legal_entities']}")
                log_callback(f"  Физлиц: {stats['physical_persons']}")
                log_callback("="*60)
            
            message = f"✓ Обработано {stats['processed']} из {stats['total_pdfs']} PDF. Результат сохранён в {os.path.basename(excel_path)}"
            return True, message, stats
            
        except Exception as e:
            self.logger.error(f"Ошибка обработки ЭЗП: {e}")
            return False, f"Ошибка: {str(e)}", stats
    
    def _detect_entity_type_offline(self, address: str) -> int:
        """
        Определяет тип получателя БЕЗ использования AI (offline mode).
        Использует EntityDetector для pattern-based классификации.
        
        Args:
            address: Адрес с названием организации или ФИО
            
        Returns:
            1 - юридическое лицо, 0 - физическое лицо
        """
        return self.entity_detector.detect_entity_type(address)
    
    def _extract_recipient_name(self, address: str) -> str:
        """
        Извлекает имя получателя из адреса
        Обычно это текст в конце адреса после запятой
        
        Args:
            address: Полный адрес
            
        Returns:
            Имя получателя (ФИО или название организации)
        """
        # Разбиваем по запятым и берём последнюю часть
        parts = [p.strip() for p in address.split(',')]
        
        # Ищем часть, которая похожа на имя (содержит заглавные буквы, но не цифры адреса)
        for part in reversed(parts):
            # Пропускаем части с цифрами (это адрес)
            if re.search(r'\d', part):
                continue
            # Если есть заглавные буквы, это может быть имя
            if re.search(r'[А-ЯЁ]', part):
                return part
        
        # Если ничего не нашли, берём последнюю часть
        return parts[-1] if parts else ""
    
    def _find_numbered_pdfs(self, folder: str) -> Dict[int, str]:
        """
        Находит нумерованные PDF файлы (1.pdf, 2.pdf, ...)
        
        Returns:
            Словарь {номер: путь к файлу}
        """
        pdf_files = {}
        
        if not os.path.exists(folder):
            return pdf_files
        
        for filename in os.listdir(folder):
            if filename.lower().endswith('.pdf'):
                # Проверяем, что имя файла - это число
                name_without_ext = filename[:-4]
                if name_without_ext.isdigit():
                    num = int(name_without_ext)
                    pdf_files[num] = os.path.join(folder, filename)
                    self.logger.debug(f"Найден PDF #{num}: {filename}")
        
        return pdf_files
    
    def validate_excel_template(self, excel_path: str) -> Tuple[bool, str]:
        """
        Проверяет корректность Excel шаблона
        
        Returns:
            (корректен, сообщение)
        """
        # CRITICAL FIX: Check openpyxl availability before using
        if not OPENPYXL_AVAILABLE:
            return False, "Библиотека openpyxl не установлена. Установите её командой: pip install openpyxl"
        
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Проверяем наличие данных
            if ws.max_row < 2:
                return False, "Excel файл пустой или содержит только заголовки"
            
            # Определяем последнюю заполненную строку
            last_filled_row = 0
            for row in range(1, ws.max_row + 1):
                has_data = False
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=row, column=col).value:
                        has_data = True
                        break
                if has_data:
                    last_filled_row = row
            
            # Устанавливаем область печати
            if last_filled_row > 0:
                print_area = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{last_filled_row}"
                ws.print_area = print_area
                self.logger.info(f"Установлена область печати: {print_area}")
            
            return True, f"Шаблон корректен. Строк с данными: {last_filled_row}"
            
        except Exception as e:
            return False, f"Ошибка проверки шаблона: {str(e)}"